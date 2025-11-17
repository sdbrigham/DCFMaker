"""
Export Handler
Handles Excel and CSV export of financial statements and DCF results
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
import re
from typing import Dict, Optional

class ExportHandler:
    """Handle exports to Excel and CSV formats"""
    
    def __init__(self, operating_model_data: Dict, dcf_results: Dict, company_name: str = "Company"):
        """
        Initialize export handler
        
        Args:
            operating_model_data: Dict with financial statements
            dcf_results: Dict with DCF calculation results
            company_name: Name of the company
        """
        self.operating_model_data = operating_model_data
        self.dcf_results = dcf_results
        # Sanitize company name for filenames
        self.company_name = re.sub(r'[<>:"/\\|?*]', '_', company_name)
    
    def format_number(self, value, format_type='millions'):
        """Format number for display"""
        if pd.isna(value) or value == 0:
            return 0.0
        
        if format_type == 'millions':
            return value / 1_000_000
        elif format_type == 'thousands':
            return value / 1_000
        else:
            return value
    
    def create_excel_workbook(self) -> BytesIO:
        """
        Create Excel workbook with multiple sheets:
        - Income Statement
        - Balance Sheet
        - Cash Flow Statement
        - DCF Summary
        """
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets
        self._create_income_statement_sheet(wb)
        self._create_balance_sheet_sheet(wb)
        self._create_cash_flow_sheet(wb)
        self._create_dcf_summary_sheet(wb)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def _create_income_statement_sheet(self, wb: Workbook):
        """Create Income Statement sheet formatted exactly like the example Excel file"""
        ws = wb.create_sheet("Historical IS")
        
        income_data = pd.DataFrame(self.operating_model_data.get('income_statement', {}))
        if income_data.empty:
            ws['B2'] = "No data available"
            return
        
        # Define colors from the example
        dark_blue = "FF002060"  # Dark blue fill and text color
        white_text = "FFFFFFFF"  # White text
        
        # Define fonts
        company_name_font = Font(bold=True, size=20, color=dark_blue)
        title_font = Font(bold=True, size=15, color=white_text)
        subtitle_font = Font(size=11, color=white_text)
        header_font = Font(bold=True, size=11, color=white_text)
        bold_font = Font(bold=True, size=11)
        regular_font = Font(size=11)
        
        # Define fills
        dark_blue_fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
        
        # Define borders
        thin_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        # Outer border for the box (C7 to I24)
        outer_border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        # Top border only (for lines above bold items)
        top_thin = Border(top=Side(style='thin'))
        # Gray fill for outer cells
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Number formats from the example
        number_format_main = '#,##0.0_);\\(#,##0.0\\)'  # Main format with parentheses for negatives
        number_format_simple = '0.0'  # For Other Operating Expenses
        number_format_dash = '\\-'  # For zero R&D values
        year_format = '####"A"'  # For year headers (2019A, 2020A, etc.)
        
        # Row 2: Company name (B2)
        ws['B2'] = self.company_name
        ws['B2'].font = company_name_font
        ws.row_dimensions[2].height = 25.8
        
        # Row 7: Title "Income Statement" (C7) with dark blue background spanning multiple columns
        # Row 8: Subtitle "$ in Millions" (C8) and year headers (G8, H8, I8, J8, etc.)
        # The example shows the title starts at C7 and spans to at least J7
        # Years start at G8 (column 7)
        
        # Define the box boundaries (content area)
        box_start_row = 7  # Content starts at row 7
        box_end_row = 24  # Based on 14 line items starting at row 11
        box_start_col = 3  # Column C (content starts here)
        
        # Calculate actual end column based on number of years
        years = sorted(income_data.columns)
        year_start_col = 7  # Column G
        actual_box_end_col = year_start_col + len(years) - 1  # Last year column
        
        # Border boundaries (one cell out from content box)
        border_start_row = box_start_row - 1  # Row 6
        border_end_row = box_end_row + 1  # Row 25
        border_start_col = box_start_col - 1  # Column B
        border_end_col = actual_box_end_col + 1  # One column after last year
        
        # Set default sheet background to gray (for all cells)
        # This will make all cells gray by default, then we'll override with white for the box
        ws.sheet_properties.tabColor = "D3D3D3"  # This doesn't set cell fill, but we'll handle it below
        
        # Title row (row 7) - inside the box, white background with dark blue header
        title_start_col = 3  # Column C
        for col in range(title_start_col, actual_box_end_col + 1):
            cell = ws.cell(row=7, column=col)
            cell.fill = dark_blue_fill
            if col == title_start_col:
                cell.value = "Income Statement"
                cell.font = title_font
        ws.row_dimensions[7].height = 19.8
        
        # Subtitle and year headers row (row 8) - inside the box, white background with dark blue header
        ws.cell(row=8, column=title_start_col).value = "$ in Millions"
        ws.cell(row=8, column=title_start_col).font = subtitle_font
        ws.cell(row=8, column=title_start_col).fill = dark_blue_fill
        
        # Fill the rest of the subtitle row with dark blue
        for col in range(title_start_col + 1, actual_box_end_col + 1):
            cell = ws.cell(row=8, column=col)
            cell.fill = dark_blue_fill
        
        # Year headers start at column G (7)
        year_start_col = 7  # Column G
        for col_idx, year in enumerate(years):
            col = year_start_col + col_idx
            cell = ws.cell(row=8, column=col)
            # Store as integer and apply format - the format ####\"A\" will display as 2023A
            cell.value = int(year)
            cell.number_format = year_format
            cell.font = header_font
            cell.fill = dark_blue_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[8].height = 15.0
        
        # Fill rows 9-10 with white background (empty rows between header and data, inside the box)
        for r in [9, 10]:
            for c in range(box_start_col, actual_box_end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Line items start at row 11, column D (4)
        line_item_col = 4  # Column D
        data_start_col = year_start_col  # Column G (7)
        
        # Line items in order (matching the example exactly)
        # Format: (label, key, is_bold, has_top_border, num_format)
        # Top border only for: Gross Profit, Operating Income, EBT, Net Income
        line_items = [
            ('Revenue', 'Revenue', True, False, number_format_main),  # Bold, no top border
            ('COGS', 'COGS', False, False, number_format_main),  # Not bold, no top border
            ('Gross Profit', 'GrossProfit', True, True, number_format_main),  # Bold, top border
            ('SG&A', 'SG&A', False, False, number_format_main),  # Not bold, no top border
            ('R&D', 'R&D', False, False, number_format_dash),  # Special format for zeros
            ('D&A', 'D&A', False, False, number_format_main),
            ('Other Operating Expenses/Income', 'OtherOperatingExpenses', False, False, number_format_simple),  # Simple format
            ('Operating Income', 'OperatingIncome', True, True, number_format_main),  # Bold, top border
            ('Other Income/(Expense), Net', 'OtherIncomeExpenseNet', False, False, number_format_main),
            ('Other Unusual Items', 'OtherUnusualItems', False, False, number_format_main),
            ('EBT', 'EBT', True, True, number_format_main),  # Bold, top border
            ('Taxes', 'TaxExpense', False, False, number_format_main),
            ('Minority Interest in Earnings', 'MinorityInterest', False, False, number_format_main),
            ('Net Income', 'NetIncome', True, True, number_format_main)  # Bold, top border
        ]
        
        # Calculate actual box end row based on number of line items
        box_end_row = 11 + len(line_items) - 1  # Start at 11, add number of items minus 1
        
        # Write line items starting at row 11
        row = 11
        for label, key, is_bold, has_top_border, num_format in line_items:
            # Write label in column D (4)
            label_cell = ws.cell(row=row, column=line_item_col)
            label_cell.value = label
            label_cell.font = bold_font if is_bold else regular_font
            # Indent non-bold items
            if not is_bold:
                label_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            else:
                label_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Set white background for cells inside the box (except header rows 7-8)
            label_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            # Apply borders: NO left border (user requested no line left of line items)
            # Top border if specified, bottom border on last row
            label_border_parts = {}
            if has_top_border:
                label_border_parts['top'] = top_thin.top
            if row == box_end_row:
                label_border_parts['bottom'] = outer_border.bottom
            label_cell.border = Border(**label_border_parts) if label_border_parts else Border()
            
            # Also fill cells in columns C, E, F with white (they're inside the box but not written to)
            for fill_col in [3, 5, 6]:  # Columns C, E, F
                if fill_col < box_start_col or fill_col > actual_box_end_col:
                    continue
                fill_cell = ws.cell(row=row, column=fill_col)
                fill_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                # Apply borders: top border if needed, bottom border on last row
                fill_border_parts = {}
                if has_top_border:
                    fill_border_parts['top'] = top_thin.top
                if row == box_end_row:
                    fill_border_parts['bottom'] = outer_border.bottom
                fill_cell.border = Border(**fill_border_parts) if fill_border_parts else Border()
            
            # Apply top border to all cells in the row if this row has top border (spanning C to actual_box_end_col)
            if has_top_border:
                for border_col in range(box_start_col, actual_box_end_col + 1):
                    border_cell = ws.cell(row=row, column=border_col)
                    # Get existing border and add top
                    existing_border = border_cell.border
                    border_cell.border = Border(
                        left=existing_border.left if existing_border.left else None,
                        top=top_thin.top,
                        right=existing_border.right if existing_border.right else None,
                        bottom=existing_border.bottom if existing_border.bottom else None
                    )
            
            # Write values for each year
            if key in income_data.index:
                for col_idx, year in enumerate(years):
                    col = data_start_col + col_idx
                    value = income_data.loc[key, year]  # key is row (line item), year is column
                    # Convert to millions
                    value_millions = self.format_number(value, 'millions')
                    
                    cell = ws.cell(row=row, column=col)
                    # For R&D, if value is exactly 0 or very close to 0, use dash format
                    # Otherwise use the main number format
                    if key == 'R&D':
                        if abs(value_millions) < 0.01:
                            cell.value = 0  # Will display as dash with the format
                            cell.number_format = number_format_dash
                        else:
                            cell.value = value_millions
                            cell.number_format = number_format_main  # Use main format for non-zero R&D
                    else:
                        cell.value = value_millions
                        cell.number_format = num_format
                    
                    cell.font = bold_font if is_bold else regular_font
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # Set white background for cells inside the box
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    
                    # Apply borders: top border if specified, NO right border (user requested no line right of years)
                    # Bottom border on last row
                    cell_border_parts = {}
                    if has_top_border:
                        cell_border_parts['top'] = top_thin.top
                    if row == box_end_row:
                        cell_border_parts['bottom'] = outer_border.bottom
                    cell.border = Border(**cell_border_parts) if cell_border_parts else Border()
            else:
                # If key not found, write zeros or dashes
                for col_idx in range(len(years)):
                    col = data_start_col + col_idx
                    cell = ws.cell(row=row, column=col)
                    if key == 'R&D':
                        cell.value = 0
                        cell.number_format = number_format_dash
                    else:
                        cell.value = 0.0
                        cell.number_format = num_format
                    cell.font = bold_font if is_bold else regular_font
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # Set white background for cells inside the box
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    
                    # Apply borders: top border if specified, NO right border, bottom border on last row
                    cell_border_parts = {}
                    if has_top_border:
                        cell_border_parts['top'] = top_thin.top
                    if row == box_end_row:
                        cell_border_parts['bottom'] = outer_border.bottom
                    cell.border = Border(**cell_border_parts) if cell_border_parts else Border()
            
            row += 1
        
        # Draw outer border around the box (one cell out from content: B6 to border_end_col+1, border_end_row+1)
        # Top border (row 6)
        for c in range(border_start_col, border_end_col + 1):
            cell = ws.cell(row=border_start_row, column=c)
            if c == border_start_col:
                cell.border = Border(left=outer_border.left, top=outer_border.top, right=None, bottom=None)
            elif c == border_end_col:
                cell.border = Border(left=None, top=outer_border.top, right=outer_border.right, bottom=None)
            else:
                cell.border = Border(left=None, top=outer_border.top, right=None, bottom=None)
        
        # Bottom border (row border_end_row)
        for c in range(border_start_col, border_end_col + 1):
            cell = ws.cell(row=border_end_row, column=c)
            if c == border_start_col:
                cell.border = Border(left=outer_border.left, top=None, right=None, bottom=outer_border.bottom)
            elif c == border_end_col:
                cell.border = Border(left=None, top=None, right=outer_border.right, bottom=outer_border.bottom)
            else:
                cell.border = Border(left=None, top=None, right=None, bottom=outer_border.bottom)
        
        # Left border (column B)
        for r in range(border_start_row + 1, border_end_row):
            cell = ws.cell(row=r, column=border_start_col)
            existing_border = cell.border
            cell.border = Border(
                left=outer_border.left,
                top=existing_border.top if existing_border.top else None,
                right=None,
                bottom=existing_border.bottom if existing_border.bottom else None
            )
        
        # Right border (column border_end_col)
        for r in range(border_start_row + 1, border_end_row):
            cell = ws.cell(row=r, column=border_end_col)
            existing_border = cell.border
            cell.border = Border(
                left=None,
                top=existing_border.top if existing_border.top else None,
                right=outer_border.right,
                bottom=existing_border.bottom if existing_border.bottom else None
            )
        
        # Fill buffer cells (one cell around the box) with white
        # Top buffer row (row 6)
        for c in range(border_start_col, border_end_col + 1):
            cell = ws.cell(row=border_start_row, column=c)
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Bottom buffer row (row border_end_row)
        for c in range(border_start_col, border_end_col + 1):
            cell = ws.cell(row=border_end_row, column=c)
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Left buffer column (column B)
        for r in range(border_start_row + 1, border_end_row):
            cell = ws.cell(row=r, column=border_start_col)
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Right buffer column (column border_end_col)
        for r in range(border_start_row + 1, border_end_row):
            cell = ws.cell(row=r, column=border_end_col)
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Fill all other cells with gray (standardize gray outside the model)
        # Fill at least 200 rows down and 100 columns to the right from the border
        max_row_to_fill = max(200, border_end_row + 200)  # At least 200 rows down
        max_col_to_fill = max(100, border_end_col + 100)  # At least 100 columns to the right
        
        for r in range(1, max_row_to_fill + 1):
            for c in range(1, max_col_to_fill + 1):
                # Skip cells inside the border area (rows border_start_row to border_end_row, columns border_start_col to border_end_col)
                if (border_start_row <= r <= border_end_row and border_start_col <= c <= border_end_col):
                    continue
                # Fill with gray (only if not already filled with something else)
                cell = ws.cell(row=r, column=c)
                current_fill = cell.fill.start_color.rgb if cell.fill and hasattr(cell.fill, 'start_color') else None
                if current_fill in [None, '00000000', 'FFFFFFFF']:  # Only fill if empty or white
                    cell.fill = gray_fill
                # Remove borders from outer cells
                cell.border = Border()
        
        # Adjust column widths (matching the example)
        ws.column_dimensions['A'].width = 13.0
        ws.column_dimensions['B'].width = 5.44
        ws.column_dimensions['C'].width = 12.44
        ws.column_dimensions['D'].width = 8.44
        ws.column_dimensions['E'].width = 13.0
        ws.column_dimensions['F'].width = 17.44
        # Year columns (G, H, I, J, etc.)
        for col_idx in range(len(years)):
            col_letter = get_column_letter(data_start_col + col_idx)
            ws.column_dimensions[col_letter].width = 13.0
    
    def _create_balance_sheet_sheet(self, wb: Workbook):
        """Create Balance Sheet sheet"""
        ws = wb.create_sheet("Balance Sheet")
        
        balance_data = pd.DataFrame(self.operating_model_data.get('balance_sheet', {}))
        if balance_data.empty:
            ws['A1'] = "No data available"
            return
        
        # Headers
        headers = ['Line Item'] + list(balance_data.index)
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Balance Sheet line items
        line_items = [
            ('Cash', 'Cash'),
            ('Short Term Investments', 'ShortTermInvestments'),
            ('Current Assets', 'CurrentAssets'),
            ('PPE', 'PPE'),
            ('Other Long Term Assets', 'OtherLongTermAssets'),
            ('Total Assets', 'TotalAssets'),
            ('Short Term Liabilities', 'ShortTermLiabilities'),
            ('Long Term Debt', 'LongTermDebt'),
            ('Long Term Leases', 'LongTermLeases'),
            ('Other Long Term Liabilities', 'OtherLongTermLiabilities'),
            ('Total Liabilities', 'TotalLiabilities'),
            ('Retained Earnings', 'RetainedEarnings'),
            ('Common Stock', 'CommonStock'),
            ('Paid-in Capital', 'PaidInCapital'),
            ('Total Equity', 'TotalEquity')
        ]
        
        row = 2
        for label, key in line_items:
            if key in balance_data.columns:
                ws.cell(row=row, column=1, value=label)
                col = 2
                for year in balance_data.index:
                    value = self.format_number(balance_data.loc[year, key])
                    ws.cell(row=row, column=col, value=value)
                    col += 1
                row += 1
        
        # Format columns
        for col in range(2, len(headers) + 1):
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.number_format = '#,##0.00'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_cash_flow_sheet(self, wb: Workbook):
        """Create Cash Flow Statement sheet"""
        ws = wb.create_sheet("Cash Flow Statement")
        
        cashflow_data = pd.DataFrame(self.operating_model_data.get('cash_flow', {}))
        if cashflow_data.empty:
            ws['A1'] = "No data available"
            return
        
        # Headers
        headers = ['Line Item'] + list(cashflow_data.index)
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Cash Flow line items
        line_items = [
            ('Net Income', 'NetIncome'),
            ('D&A', 'D&A'),
            ('Change in Working Capital', 'ChangeInWorkingCapital'),
            ('Operating Cash Flow', 'OperatingCashFlow'),
            ('Capital Expenditures', 'CapitalExpenditures'),
            ('Investing Cash Flow', 'InvestingCashFlow'),
            ('Financing Cash Flow', 'FinancingCashFlow'),
            ('Net Cash Flow', 'NetCashFlow')
        ]
        
        row = 2
        for label, key in line_items:
            if key in cashflow_data.columns:
                ws.cell(row=row, column=1, value=label)
                col = 2
                for year in cashflow_data.index:
                    value = self.format_number(cashflow_data.loc[year, key])
                    ws.cell(row=row, column=col, value=value)
                    col += 1
                row += 1
        
        # Format columns
        for col in range(2, len(headers) + 1):
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.number_format = '#,##0.00'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_dcf_summary_sheet(self, wb: Workbook):
        """Create DCF Summary sheet"""
        ws = wb.create_sheet("DCF Summary")
        
        # Title
        ws['A1'] = "DCF Valuation Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        row = 3
        
        # Assumptions
        ws.cell(row=row, column=1, value="Assumptions").font = Font(bold=True, size=12)
        row += 1
        
        assumptions = self.dcf_results.get('assumptions', {})
        assumption_labels = {
            'risk_free_rate': 'Risk-Free Rate',
            'beta': 'Beta',
            'market_risk_premium': 'Market Risk Premium',
            'cost_of_debt': 'Cost of Debt',
            'tax_rate': 'Tax Rate',
            'debt_to_equity': 'Debt-to-Equity Ratio',
            'terminal_growth_rate': 'Terminal Growth Rate'
        }
        
        for key, label in assumption_labels.items():
            value = assumptions.get(key, 'N/A')
            ws.cell(row=row, column=1, value=label)
            if isinstance(value, (int, float)):
                ws.cell(row=row, column=2, value=value).number_format = '0.00%' if key != 'beta' else '0.00'
            else:
                ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
        
        # WACC
        ws.cell(row=row, column=1, value="WACC").font = Font(bold=True)
        wacc = self.dcf_results.get('wacc', 0)
        ws.cell(row=row, column=2, value=wacc).number_format = '0.00%'
        row += 2
        
        # Free Cash Flows
        ws.cell(row=row, column=1, value="Free Cash Flows (in millions)").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value="Year")
        ws.cell(row=row, column=2, value="FCF")
        row += 1
        
        fcf = self.dcf_results.get('free_cash_flows', {})
        for year, value in sorted(fcf.items()):
            ws.cell(row=row, column=1, value=str(year))
            ws.cell(row=row, column=2, value=self.format_number(value)).number_format = '#,##0.00'
            row += 1
        
        row += 1
        
        # Terminal Value
        ws.cell(row=row, column=1, value="Terminal Value (in millions)").font = Font(bold=True)
        terminal_value = self.dcf_results.get('terminal_value', 0)
        ws.cell(row=row, column=2, value=self.format_number(terminal_value)).number_format = '#,##0.00'
        row += 2
        
        # Valuation Summary
        ws.cell(row=row, column=1, value="Valuation Summary (in millions)").font = Font(bold=True, size=12)
        row += 1
        
        pv_fcf = self.dcf_results.get('total_pv_fcf', 0)
        pv_terminal = self.dcf_results.get('present_value_terminal', 0)
        enterprise_value = self.dcf_results.get('enterprise_value', 0)
        equity_value = self.dcf_results.get('equity_value', 0)
        
        ws.cell(row=row, column=1, value="PV of FCFs")
        ws.cell(row=row, column=2, value=self.format_number(pv_fcf)).number_format = '#,##0.00'
        row += 1
        
        ws.cell(row=row, column=1, value="PV of Terminal Value")
        ws.cell(row=row, column=2, value=self.format_number(pv_terminal)).number_format = '#,##0.00'
        row += 1
        
        ws.cell(row=row, column=1, value="Enterprise Value").font = Font(bold=True)
        ws.cell(row=row, column=2, value=self.format_number(enterprise_value)).number_format = '#,##0.00'
        row += 1
        
        ws.cell(row=row, column=1, value="Equity Value").font = Font(bold=True)
        ws.cell(row=row, column=2, value=self.format_number(equity_value)).number_format = '#,##0.00'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def export_to_csv(self, output_dir: str = ".") -> Dict[str, str]:
        """
        Export financial statements to separate CSV files
        
        Returns:
            Dict with file paths
        """
        files = {}
        
        # Income Statement
        income_data = pd.DataFrame(self.operating_model_data.get('income_statement', {}))
        if not income_data.empty:
            income_data_formatted = income_data / 1_000_000  # Convert to millions
            filepath = os.path.join(output_dir, f"{self.company_name}_Income_Statement.csv")
            income_data_formatted.to_csv(filepath)
            files['income_statement'] = filepath
        
        # Balance Sheet
        balance_data = pd.DataFrame(self.operating_model_data.get('balance_sheet', {}))
        if not balance_data.empty:
            balance_data_formatted = balance_data / 1_000_000  # Convert to millions
            filepath = os.path.join(output_dir, f"{self.company_name}_Balance_Sheet.csv")
            balance_data_formatted.to_csv(filepath)
            files['balance_sheet'] = filepath
        
        # Cash Flow
        cashflow_data = pd.DataFrame(self.operating_model_data.get('cash_flow', {}))
        if not cashflow_data.empty:
            cashflow_data_formatted = cashflow_data / 1_000_000  # Convert to millions
            filepath = os.path.join(output_dir, f"{self.company_name}_Cash_Flow.csv")
            cashflow_data_formatted.to_csv(filepath)
            files['cash_flow'] = filepath
        
        # DCF Summary
        dcf_summary = pd.DataFrame({
            'Metric': ['WACC', 'Terminal Value', 'PV of FCFs', 'PV of Terminal Value', 
                      'Enterprise Value', 'Equity Value'],
            'Value (millions)': [
                self.dcf_results.get('wacc', 0) * 100,  # Convert to percentage
                self.format_number(self.dcf_results.get('terminal_value', 0)),
                self.format_number(self.dcf_results.get('total_pv_fcf', 0)),
                self.format_number(self.dcf_results.get('present_value_terminal', 0)),
                self.format_number(self.dcf_results.get('enterprise_value', 0)),
                self.format_number(self.dcf_results.get('equity_value', 0))
            ]
        })
        filepath = os.path.join(output_dir, f"{self.company_name}_DCF_Summary.csv")
        dcf_summary.to_csv(filepath, index=False)
        files['dcf_summary'] = filepath
        
        return files

